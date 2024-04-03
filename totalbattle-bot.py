# TotalBattle.bot
# Copyright (c) 2024 Grigore Stefan <g_stefan@yahoo.com.hpp>
# Apache License 2.0 <https://opensource.org/license/apache-2-0/>
# SPDX-FileCopyrightText: 2024 Grigore Stefan <g_stefan@yahoo.com.hpp>
# SPDX-License-Identifier: Apache-2.0

#
# Version 1.2.0 2024-03-31
#

import flet
from flet import Page, Text, TextField, Switch, ElevatedButton, Row, Divider, DatePicker, TimePicker, icons, SnackBar
import threading
import cv2 as cv
import numpy as np
import pyautogui
import time
from datetime import datetime, timedelta
import csv
import os
import os.path
import pytesseract
import easyocr
import xlsxwriter
import openpyxl


# ---
ocrProcessor = easyocr.Reader(["en","de","fr"])
useEasyOCR = False
ocrMemoryImageFrom = {}
ocrMemoryImageName = {}
ocrMemoryImageSource = {}
ocrMemoryImageContains = {}
# ---
buttonDeleteImage = cv.imread("images/button-delete.png", cv.IMREAD_GRAYSCALE)
buttonGiftsActiveImage = cv.imread("images/button-gifts-active.png", cv.IMREAD_GRAYSCALE)
buttonOpenImage = cv.imread("images/button-open.png", cv.IMREAD_GRAYSCALE)
windowClanTitleImage = cv.imread("images/window-clan-title.png", cv.IMREAD_GRAYSCALE)

buttonDeleteImageX = []
buttonGiftsActiveImageX = []
buttonOpenImageX = []
windowClanTitleImageX = []

# ---
screenShotRGB = None
screenShotGray = None
cropImage = None
screenShotZoom = 1
screenShotZoomList = []
screenShotZoomIndex = 0
# ---
topX=0
topY=0
posXMouseDelta=0
posYMouseDelta=0
posXMouse=0
posYMouse=0
# ---
giftsCaptureRegionX=0
giftsCaptureRegionY=0
giftsCaptureRegionLnX=0
giftsCaptureRegionLnY=0
# ---
giftNameRegionX=0
giftNameRegionY=0
giftNameRegionLnX=0
giftNameRegionLnY=0
# ---
giftFromRegionX=0
giftFromRegionY=0
giftFromRegionLnX=0
giftFromRegionLnY=0
# ---
giftSourceRegionX=0
giftSourceRegionY=0
giftSourceRegionLnX=0
giftSourceRegionLnY=0
# ---
giftContainsRegionX=0
giftContainsRegionY=0
giftContainsRegionLnX=0
giftContainsRegionLnY=0
# ---
threadStarted=False
stopProcessing=False
processingDone=False

def getScreenshot():
    global screenShotRGB,screenShotGray,screenShotZoom
    screenShotRGB=cv.cvtColor(np.array(pyautogui.screenshot()),cv.COLOR_RGB2BGR)
    newLnX = screenShotRGB.shape[1]/screenShotZoom
    newLnY = screenShotRGB.shape[0]/screenShotZoom    
    screenShotRGB = cv.resize(screenShotRGB, (int(newLnX), int(newLnY)), fx=0, fy=0, interpolation = cv.INTER_AREA)
    screenShotGray=cv.cvtColor(screenShotRGB,cv.COLOR_RGB2GRAY)

def getScreenshotX(x,y,lnX,lnY):
    global screenShotRGB,screenShotGray,screenShotZoom
    screenShotRGB=cv.cvtColor(np.array(pyautogui.screenshot(region=(int(x*screenShotZoom),int(y*screenShotZoom),int(lnX*screenShotZoom),int(lnY*screenShotZoom)))),cv.COLOR_RGB2BGR)
    screenShotRGB = cv.resize(screenShotRGB, (int(screenShotRGB.shape[1]/screenShotZoom), int(screenShotRGB.shape[0]/screenShotZoom)), fx=0, fy=0, interpolation = cv.INTER_AREA)
    screenShotGray=cv.cvtColor(screenShotRGB,cv.COLOR_RGB2GRAY)

def matchImage(image):
    global screenShotGray,screenShotRGB,topX,topY,posXMouse,posYMouse
    w, h = image.shape[::-1]
    res = cv.matchTemplate(screenShotGray,image,cv.TM_CCOEFF_NORMED)
    loc = np.where( res >= 0.8)    
    found = False
    for pt in zip(*loc[::-1]):        
        topX=int(pt[0])
        topY=int(pt[1])
        posXMouse=int(pt[0] + w/2)
        posYMouse=int(pt[1] + h/2)
        cv.rectangle(screenShotRGB, pt, (pt[0] + w, pt[1] + h), (0,0,255), 2)
        found = True
        break    
    return found

def saveGiftsCaptureRegion():
    global giftsCaptureRegionX,giftsCaptureRegionY,giftsCaptureRegionLnX,giftsCaptureRegionLnY
    global giftNameRegionX,giftNameRegionY,giftNameRegionLnX,giftNameRegionLnY
    global giftFromRegionX,giftFromRegionY,giftFromRegionLnX,giftFromRegionLnY
    global giftSourceRegionX,giftSourceRegionY,giftSourceRegionLnX,giftSourceRegionLnY
    global giftContainsRegionX,giftContainsRegionY,giftContainsRegionLnX,giftContainsRegionLnY

    global posXMouseDelta,posYMouseDelta
    # ---
    giftsCaptureRegionX=topX-36
    giftsCaptureRegionY=topY+110
    giftsCaptureRegionLnX=625
    giftsCaptureRegionLnY=94
    # --- 
    giftNameRegionX=4
    giftNameRegionY=4
    giftNameRegionLnX=404
    giftNameRegionLnY=31
    # ---
    giftFromRegionX=48
    giftFromRegionY=37
    giftFromRegionLnX=392
    giftFromRegionLnY=21
    # ---
    giftSourceRegionX=57
    giftSourceRegionY=58
    giftSourceRegionLnX=375
    giftSourceRegionLnY=20
    # ---
    giftContainsRegionX=70
    giftContainsRegionY=77
    giftContainsRegionLnX=421
    giftContainsRegionLnY=17
    # ---
    posXMouseDelta=giftsCaptureRegionX
    posYMouseDelta=giftsCaptureRegionY
    # ---

def getGiftScreenshot():
    global giftsCaptureRegionX,giftsCaptureRegionY,giftsCaptureRegionLnX,giftsCaptureRegionLnY
    getScreenshotX(giftsCaptureRegionX,giftsCaptureRegionY,giftsCaptureRegionLnX,giftsCaptureRegionLnY)

def saveScreenshot(info):
    global screenShotRGB
    cv.imwrite(info+"-screenshot.png",screenShotRGB)    

def saveCropImage(info):
    global cropImage
    cv.imwrite(info+"-crop.png",cropImage)

def clickX():
    global posXMouse,posYMouse,posXMouseDelta,posYMouseDelta,screenShotZoom
    pyautogui.moveTo(int((posXMouse+posXMouseDelta)*screenShotZoom),int((posYMouse+posYMouseDelta)*screenShotZoom),0.2)
    pyautogui.click(int((posXMouse+posXMouseDelta)*screenShotZoom),int((posYMouse+posYMouseDelta)*screenShotZoom))

def clickAt(x,y):
    global screenShotZoom
    pyautogui.moveTo(int((posXMouse+posXMouseDelta)*screenShotZoom),int((posYMouse+posYMouseDelta)*screenShotZoom),0.2)
    pyautogui.click(int(x),int(y))

def ocrEasyOCR(processedImage):
    result = ocrProcessor.readtext(processedImage)
    text = ""
    for (bbox, textFound, prob) in result:
        if prob > 0.8:
            text += textFound + " "
    return text.strip()

def ocr(image,x,y,lnX,lnY,imageMemory):
    global cropImage,ocrProcessor,useEasyOCR    
    cropImage = image[y:y+lnY, x:x+lnX]

    fastImage = cv.resize(cropImage, (int(cropImage.shape[1]/2), int(cropImage.shape[0]/2)), fx=0, fy=0, interpolation = cv.INTER_AREA)
    #cv.imwrite(datetime.now().strftime("%Y-%m-%d_%H-%M-%S_")+"-fast.png",fastImage)
    
    for processed in imageMemory:
        res = cv.matchTemplate(fastImage,imageMemory[processed],cv.TM_CCOEFF_NORMED)
        loc = np.where( res >= 0.9)
        for pt in zip(*loc[::-1]):
            return processed
        
    cropImage = cv.resize(cropImage, (cropImage.shape[1]*4, cropImage.shape[0]*4), fx=0, fy=0, interpolation = cv.INTER_AREA)
    normImage = np.zeros((cropImage.shape[0], cropImage.shape[1]))
    cropImage = cv.normalize(cropImage, normImage, 0, 255, cv.NORM_MINMAX)
    cropImage = cv.threshold(cropImage, 100, 255, cv.THRESH_BINARY)[1]
    cropImage = cv.medianBlur(cropImage, 3)

    if useEasyOCR:
        text = ocrEasyOCR(cropImage)
        if not text == "":
            imageMemory[text]=fastImage
            return text

    text = pytesseract.image_to_string(cropImage, config="--psm 12 --oem 1")
    text = text.strip()
    if not text == "":
        imageMemory[text]=fastImage
        return text

    text = ocrEasyOCR(cropImage)
    if not text == "":
        imageMemory[text]=fastImage
            
    return text

def getGiftName():
    global screenShotGray
    global giftNameRegionX,giftNameRegionY,giftNameRegionLnX,giftNameRegionLnY
    global ocrMemoryImageName
    text = ocr(screenShotGray,giftNameRegionX,giftNameRegionY,giftNameRegionLnX,giftNameRegionLnY,ocrMemoryImageName)
    text = text.strip()
    return text

def getGiftFrom():
    global screenShotGray
    global giftFromRegionX,giftFromRegionY,giftFromRegionLnX,giftFromRegionLnY
    global ocrMemoryImageFrom
    text = ocr(screenShotGray,giftFromRegionX,giftFromRegionY,giftFromRegionLnX,giftFromRegionLnY,ocrMemoryImageFrom)
    text = text.strip()
    return text

def getGiftSource():
    global screenShotGray
    global giftSourceRegionX,giftSourceRegionY,giftSourceRegionLnX,giftSourceRegionLnY
    global ocrMemoryImageSource
    text = ocr(screenShotGray,giftSourceRegionX,giftSourceRegionY,giftSourceRegionLnX,giftSourceRegionLnY,ocrMemoryImageSource)
    text = text.strip()
    return text

def getGiftContains():
    global screenShotGray
    global giftContainsRegionX,giftContainsRegionY,giftContainsRegionLnX,giftContainsRegionLnY
    global ocrMemoryImageContains
    text = ocr(screenShotGray,giftContainsRegionX,giftContainsRegionY,giftContainsRegionLnX,giftContainsRegionLnY,ocrMemoryImageContains)
    text = text.strip()
    return text

def matchImageWindowClanTitle():    
    global windowClanTitleImageX
    for image in windowClanTitleImageX:
        if matchImage(image):
            return True
    return False

def matchImageButtonGiftsActive():    
    global buttonGiftsActiveImageX
    for image in buttonGiftsActiveImageX:
        if matchImage(image):
            return True
    return False
    
def matchImageButtonDelete():
    global buttonDeleteImageX
    for image in buttonDeleteImageX:
        if matchImage(image):
            return True
    return False

def matchImageButtonOpen():
    global buttonOpenImageX
    for image in buttonOpenImageX:
        if matchImage(image):
            return True
    return False

# ---
def writeXLSX(table, header, filename):
    workbook = xlsxwriter.Workbook(filename)

    headerFormat = workbook.add_format(
        {
            "bold": True,
            "align": "center",
            "valign": "vcenter",            
            "border": 1,
        }
    )

    worksheet = workbook.add_worksheet()
    worksheet.freeze_panes(1, 0)    

    rowIndex = 0
    colIndex = 0
    for col in header:
        worksheet.write(rowIndex, colIndex, col,headerFormat)
        colIndex += 1

    rowIndex = 1    
    for row in (table):
        colIndex = 0
        for col in row:
            worksheet.write(rowIndex, colIndex, col)
            colIndex += 1
        rowIndex += 1

    workbook.close()

def readXLSX(filename,numCols):
    table=[]
    if not os.path.isfile(filename):
        return  table

    workbook = openpyxl.load_workbook(filename)    
    worksheet = workbook.active
      
    for row in range(1, worksheet.max_row):
        line=[]
        for item in range(0, numCols):
             line.append("")
        index = 0
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if not index < numCols:
                 break
            line[index]=str(col[row].value).strip()
            index = index + 1
        isEmpty = True
        for item in range(0, numCols):
            if not line[item] == "":
                isEmpty = False
                break
        if isEmpty:
            continue
        table.append(line)
        
    return table
# ---
tableOcrFixGiftContent=[]

def ocrFixGiftContentLoad():
    global tableOcrFixGiftContent
    tableOcrFixGiftContent = readXLSX("./config/ocr-fix-gift-content.xlsx",2)
    #print(tableOcrFixGiftContent)

def ocrFixGiftContent(value):
    global tableOcrFixGiftContent
    for item in tableOcrFixGiftContent:
        if item[0] == value:
            return item[1].strip()
    return value.strip()
    
# ---
tableOcrFixGiftFrom=[]

def ocrFixGiftFromLoad():
    global tableOcrFixGiftFrom
    tableOcrFixGiftFrom = readXLSX("./config/ocr-fix-gift-from.xlsx",2)
    #print(tableOcrFixGiftFrom)

def ocrFixGiftFrom(value):
    global tableOcrFixGiftFrom
    for item in tableOcrFixGiftFrom:
        if item[0] == value:
            return item[1].strip()
    return value.strip()

# ---
tableOcrFixGiftName=[]

def ocrFixGiftNameLoad():
    global tableOcrFixGiftName
    tableOcrFixGiftName = readXLSX("./config/ocr-fix-gift-name.xlsx",2)
    #print(tableOcrFixGiftName)

def ocrFixGiftName(value):
    global tableOcrFixGiftName
    for item in tableOcrFixGiftName:
        if item[0] == value:
            return item[1].strip()
    return value.strip()

# ---
tableOcrFixGiftSource=[]

def ocrFixGiftSourceLoad():
    global tableOcrFixGiftSource    
    tableOcrFixGiftSource = readXLSX("./config/ocr-fix-gift-source.xlsx",2)
    #print(tableOcrFixGiftSource)

def ocrFixGiftSource(value):
    global tableOcrFixGiftSource
    for item in tableOcrFixGiftSource:
        if item[0] == value:
            return item[1].strip()
    return value.strip()

# ---
tableGiftScore=[]

def giftScoreLoad():
    global tableGiftScore
    tableGiftScore = readXLSX("./config/gift-score.xlsx",2)
    #print(tableGiftScore)

def getGiftScore(value):
    global tableGiftScore
    for item in tableGiftScore:
        if item[0].lower() == value.lower():
            return int(item[1].strip())
    return 0

# ---
tablePlayerList=[]

def playerListLoad():
    global tablePlayerList
    tablePlayerList = readXLSX("./config/player-list.xlsx",1)
    #print(tablePlayerList)
                        
def playerListSave():
    global tablePlayerList
    writeXLSX(tablePlayerList, ["Player"], "./config/player-list.xlsx")
    
# ---
def sortSecond(val):    
    return val[1]
# ---
tableGiftIgnore=[]

def giftIgnoreLoad():
    global tableGiftIgnore
    tableGiftIgnore = readXLSX("./config/gift-ignore.xlsx",1)
    #print(tableGiftIgnore)

def giftIgnore(value):
    global tableGiftIgnore
    for item in tableGiftIgnore:
        if item[0].lower() == value.lower():
            return True
    return False

# ---
def prepareMatch(matchImage):
    matchList = []
    matchList.append(matchImage)
    matchList.append(cv.resize(matchImage, (matchImage.shape[1]+1, matchImage.shape[0]), fx=0, fy=0, interpolation = cv.INTER_AREA))
    matchList.append(cv.resize(matchImage, (matchImage.shape[1]+2, matchImage.shape[0]), fx=0, fy=0, interpolation = cv.INTER_AREA))    
    matchList.append(cv.resize(matchImage, (matchImage.shape[1], matchImage.shape[0]+1), fx=0, fy=0, interpolation = cv.INTER_AREA))
    matchList.append(cv.resize(matchImage, (matchImage.shape[1], matchImage.shape[0]+2), fx=0, fy=0, interpolation = cv.INTER_AREA))    
    matchList.append(cv.resize(matchImage, (matchImage.shape[1]+1, matchImage.shape[0]+1), fx=0, fy=0, interpolation = cv.INTER_AREA))
    matchList.append(cv.resize(matchImage, (matchImage.shape[1]+2, matchImage.shape[0]+2), fx=0, fy=0, interpolation = cv.INTER_AREA))    
    return matchList

def initProcessing():
    global buttonDeleteImage
    global buttonGiftsActiveImage
    global buttonOpenImage
    global windowClanTitleImage
    global buttonDeleteImageX
    global buttonGiftsActiveImageX
    global buttonOpenImageX
    global windowClanTitleImageX
    global screenShotZoomList,screenShotZoomIndex
    
    buttonDeleteImageX=prepareMatch(buttonDeleteImage)
    buttonGiftsActiveImageX=prepareMatch(buttonGiftsActiveImage)
    buttonOpenImageX=prepareMatch(buttonOpenImage)
    windowClanTitleImageX=prepareMatch(windowClanTitleImage)

    screenShotZoomList = []
    screenShotZoomIndex = 0
    screenShotZoomList.append(1)
    screenShotZoomList.append(1.33334)
    screenShotZoomList.append(2)
# ---

def main(page: Page):
    global stopProcessing,processingDone,threadStarted,useEasyOCR
    page.window_width = 360
    page.window_height = 600 
    page.title = "TotalBatttle.bot"
    page.vertical_alignment = "top"    

    status = Text("ready")
    stopProcessing = False
    processingDone = False    

    txtNumber = TextField(value="1000", text_align="right", width=100)
    switchUseEasyOCR = Switch(label="EasyOCR", value=useEasyOCR)

    def threadProc():
        global stopProcessing, processingDone,threadStarted
        global screenShotZoom,screenShotZoomList,screenShotZoomIndex
        global posXMouse,posYMouse,posXMouseDelta,posYMouseDelta
        threadStarted = True
        status.value = "starting up ..."        
        page.update()
        initProcessing()
        screenShotZoom = 1
        count = 0
        countLN = int(txtNumber.value)
        while not stopProcessing:
            time.sleep(100/1000)
            posXMouse = 0
            posYMouse = 0
            posXMouseDelta = 0
            posYMouseDelta = 0
            getScreenshot()
            #saveScreenshot(datetime.now().strftime("%Y-%m-%d_%H-%M-%S")+"-main-"+str(screenShotZoom))
            if not matchImageWindowClanTitle():
                screenShotZoomIndex = screenShotZoomIndex + 1
                if screenShotZoomIndex < len(screenShotZoomList):
                    screenShotZoom = screenShotZoomList[screenShotZoomIndex]
                    status.value = "search gifts window "+str(screenShotZoom)
                    page.update()
                    continue
                status.value = "no gifts window found"
                page.update()
                time.sleep(2)                
                break
            saveGiftsCaptureRegion()            
            if not matchImageButtonGiftsActive():
                status.value = "no gifts button found"
                page.update()
                time.sleep(2)               
                break
            while not stopProcessing:
                status.value = "process gift ..."
                page.update()
                getGiftScreenshot()

                dateNow = datetime.now()
                giftDate = dateNow.strftime("%Y-%m-%d %H:%M:%S")
                dateDay = dateNow.strftime("%Y-%m-%d")

                #saveScreenshot(datetime.now().strftime("%Y-%m-%d_%H-%M-%S")+"-gift")

                isDeleted = False
                if matchImageButtonDelete():
                    isDeleted = True
                if not matchImageButtonOpen():
                    if not isDeleted:                        
                        status.value = "no more gifts, done"
                        page.update()
                        time.sleep(2)
                        stopProcessing=True
                        break                                
                
                giftFrom = getGiftFrom()
                if giftFrom == "":
                    saveScreenshot("./errors/"+dateNow.strftime("%Y-%m-%d_%H-%M-%S_"))
                    saveCropImage("./errors/"+dateNow.strftime("%Y-%m-%d_%H-%M-%S_"))
                    status.value = "OCR Error - Gift From"
                    page.update()
                    stopProcessing = True
                    processingDone = True                
                    threadStarted = False
                    return                
                            
                giftName = getGiftName()
                giftSource = getGiftSource()
                giftContains = "unknown"
                giftStatus = "Ok"
                if isDeleted:
                    giftContains = getGiftContains()
                    giftStatus = "Deleted"

                #print("Gift Name: "+giftName)
                #print("Gift From: "+giftFrom)
                #print("Gift Source: "+giftSource)
                    
                fileName = "./repository/"+dateDay+"-chest-info.csv"
                outFile = open(fileName, "a", encoding='utf8')
                outFile.write("\""+giftDate+"\",\""+giftFrom+"\",\""+giftName+"\",\""+giftSource+"\",\""+giftContains+"\",\""+giftStatus+"\"\n")
                outFile.close()
                #---
                count = count + 1
                status.value = "["+str(count)+"/"+str(countLN)+"] "+giftFrom+", "+giftName
                page.update()                
                #---
                clickX()
                #print("ClickX: "+str(int(posXMouse+posXMouseDelta))+" Y:"+str(int(posYMouse+posYMouseDelta)))
                time.sleep(25/1000)
                pyautogui.move(50, 50, 25/1000)
                # ---                     
                if count >= countLN:
                    break
                # --- next
            break
        status.value = "done"
        page.update()  
        stopProcessing = True
        processingDone = True                
        threadStarted = False

    def getChests(e):
        global stopProcessing, processingDone, threadStarted,useEasyOCR
        page.update()
        if threadStarted:
            page.update()
            return
        useEasyOCR = switchUseEasyOCR.value        
        threadStarted = True
        stopProcessing = False
        processingDone = False
        status.value = "preparing"
        page.update()
        (threading.Timer(1.0,threadProc)).start()        

    def stopProcess(e):
        global stopProcessing, processingDone
        stopProcessing = True
        page.update()

        
    # ---
    dateStart = datetime.now() #(datetime.now()-timedelta(days=1))
    dateStartText = dateStart.strftime("%Y-%m-%d")

    async def changeDateStart(e):
        dateStart=datepickerStart.value        
        txtDateStart.value= dateStart.strftime("%Y-%m-%d")
        e.control.page.update()

    datepickerStart = DatePicker(value=dateStart,on_change=changeDateStart)
    txtDateStart = TextField(value=dateStartText, text_align="center", width=120)

    async def openStartDatePicker(e):
            datepickerStart.pick_date()            
    
    buttonPickStartDate = ElevatedButton(
                    "Pick start date",
                    icon=icons.CALENDAR_MONTH,
                    on_click=openStartDatePicker,
                    width=180
                )
    
    # ---
    dateEnd = datetime.now() #(datetime.now()-timedelta(days=1))
    dateEndText = dateEnd.strftime("%Y-%m-%d")

    async def changeDateEnd(e):
        dateEnd=datepickerEnd.value        
        txtDateEnd.value= dateEnd.strftime("%Y-%m-%d")
        e.control.page.update()

    datepickerEnd = DatePicker(value=dateEnd,on_change=changeDateEnd)
    txtDateEnd = TextField(value=dateEndText, text_align="center", width=120)

    async def openEndDatePicker(e):    
            datepickerEnd.pick_date()            
    
    buttonPickEndDate = ElevatedButton(
                    "Pick end date",
                    icon=icons.CALENDAR_MONTH,
                    on_click=openEndDatePicker,
                    width=180
                )
    # ---
    def cmdButtonSubmit(e):
        strDateStart=txtDateStart.value
        strDateEnd=txtDateEnd.value
        if strDateEnd < strDateStart:
            strDateStart=txtDateEnd.value
            strDateEnd=txtDateStart.value
        dateStart = datetime.strptime(strDateStart,"%Y-%m-%d")
        dateEnd = datetime.strptime(strDateEnd,"%Y-%m-%d")
        dateList = []
        while dateStart <= dateEnd:
            dateList.append(dateStart)
            dateStart += timedelta(days=1)
        giftsTable=[]
        for day in dateList:
            filename="./repository/"+day.strftime("%Y-%m-%d")+"-chest-info.csv"
            if os.path.isfile(filename):
                with open(filename, newline='', encoding='utf8') as csvFile:
                    csvReader = csv.reader(csvFile, delimiter=',', quotechar='\"',quoting=csv.QUOTE_ALL)
                    for row in csvReader:
                        if len(row) == 0:
                            continue
                        giftsTable.append(row)
        #print(giftsTable)        
        ocrFixGiftContentLoad()
        ocrFixGiftFromLoad()
        ocrFixGiftNameLoad()
        ocrFixGiftSourceLoad()
        giftScoreLoad()
        playerListLoad()
        giftIgnoreLoad()

        playerStats={}
        giftStats={}
        giftsTableFinal=[]
        for line in giftsTable:
            if len(line)==0:
                continue
            giftFrom = ocrFixGiftFrom(line[1])
            giftName = ocrFixGiftName(line[2])
            giftSource = ocrFixGiftSource(line[3])
            giftContains = ocrFixGiftContent(line[4])
            giftScore = getGiftScore(giftSource)

            giftsTableFinal.append([line[0],giftFrom,giftName,giftSource,giftContains,line[5],giftScore])
            if giftIgnore(giftSource):
               continue

            if not giftFrom in playerStats:
                playerStats[giftFrom]=["",0,0,False]
            playerStats[giftFrom][0]=giftFrom
            playerStats[giftFrom][1]+=giftScore
            playerStats[giftFrom][2]+=1

            if not giftSource in giftStats:
                giftStats[giftSource]=[giftSource,0]
            giftStats[giftSource][1]+=1

        reportDate=datetime.now().strftime("%Y-%m-%d")        
        filename="./report/"+reportDate+"-chest-info.xlsx"
        writeXLSX(giftsTableFinal, ["Date","Player","Name","Source","Content","Status","Score"], filename)
        
        # ---
        playerStatsSort = []        
        for player in playerStats:
             playerStatsSort.append([playerStats[player][0],playerStats[player][1],playerStats[player][2]])
        playerStatsSort.sort(key=sortSecond) 
        playerStatsSort.reverse()

        # ---
        playerStatsTable = []        
        for line in playerStatsSort:
             playerStatsTable.append(line)

        filename="./report/"+reportDate+"-player-top.xlsx"
        writeXLSX(playerStatsTable, ["Player","Score","Count"], filename)        

        # ---
        giftStatsSort = []
        for gift in giftStats:
            giftStatsSort.append([giftStats[gift][0],giftStats[gift][1]])
        giftStatsSort.sort(key=sortSecond) 
        giftStatsSort.reverse()

        giftsTableStats=[]        
        for line in giftStatsSort:
            giftsTableStats.append(line)
        
        filename="./report/"+reportDate+"-gifts-top.xlsx"
        writeXLSX(giftsTableStats, ["Name","Count"], filename)
        
        # ---
        global tablePlayerList

        playerInfoTable=[]        
        playerNoGifts=[]        
        for line in tablePlayerList:
            player = line[0]
            if player in playerStats:
                playerInfoTable.append([playerStats[player][0],playerStats[player][1],playerStats[player][2]])
                playerStats[player][3]=True
            else:
                playerInfoTable.append([player,0,0])
                playerNoGifts.append([player])

        for player in playerStats:
             if not playerStats[player][3]:
                  tablePlayerList.append([player])
                  playerInfoTable.append([playerStats[player][0],playerStats[player][1],playerStats[player][2]])                  

        playerListSave()

        filename="./report/"+reportDate+"-player-stats.xlsx"
        writeXLSX(playerInfoTable, ["Player","Score","Count"], filename)
        
        filename="./report/"+reportDate+"-player-no-gifts.xlsx"
        writeXLSX(playerNoGifts, ["Player"], filename)
        
        e.control.page.snack_bar = SnackBar(Text("Report done!"))
        e.control.page.snack_bar.open = True
        e.control.page.update()
        return
    
    buttonSubmit = ElevatedButton(text="Submit", on_click=cmdButtonSubmit)
    # ---
    def cmdButtonOpenFolder(e):
        os.startfile(os.path.normpath("./report/"))
        return

    buttonOpenFolder = ElevatedButton(text="Open Folder", on_click=cmdButtonOpenFolder)
    # ---

    page.add(        
        Row(
            [
                Text("Status:"),
                status,
            ],
            alignment="left",
        ),
        Row(
            [
                Text("Count:"),
                txtNumber,
            ],
            alignment="left",
        ),
        Row(
            [
                ElevatedButton(text="Get Chests", on_click=getChests),                
            ],
            alignment="center",
        ),
        Row(
            [
                ElevatedButton(text="Stop", on_click=stopProcess),
            ],
            alignment="center",
        ),
        Divider(),
        Row(
            [                
                switchUseEasyOCR
            ],
            alignment="left",
        ),
        Divider(),
        Row(
            [
                Text("Report"),
            ],
            alignment="left",
        ),
        Row(
            [
                datepickerStart,
                buttonPickStartDate,
                txtDateStart
            ],
            alignment="left",
        ),
        Row(
            [
                datepickerEnd,
                buttonPickEndDate,
                txtDateEnd
            ],
            alignment="left",
        ),
        Row(
            [
                buttonSubmit,
                buttonOpenFolder
            ],
            alignment="left",
        )        
        )


flet.app(target=main)
